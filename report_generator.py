"""
Financial KPI Dashboard — Automated Excel Report Generator
Author: Satyam Thakur
Description: Connects to MySQL, runs KPI queries, and generates
             a formatted multi-sheet Excel report automatically.
Run:  python report_generator.py
"""

import os
import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── Config ────────────────────────────────────────────────
DB_CONFIG = {
    "host":     "localhost",
    "user":     "root",
    "password": "your_password_here",    # update this
    "database": "financial_kpi_db"
}

OUTPUT_DIR  = "reports"
REPORT_NAME = f"Financial_KPI_Report_{datetime.now().strftime('%Y-%m')}.xlsx"

# ─── Color palette ─────────────────────────────────────────
NAVY   = "1A3557"
ACCENT = "2563EB"
LIGHT  = "DBEAFE"
GREEN  = "16A34A"
RED    = "DC2626"
GRAY   = "6B7280"


def get_connection():
    return mysql.connector.connect(**DB_CONFIG)


def fetch_df(query: str) -> pd.DataFrame:
    """Execute a query and return a DataFrame."""
    conn = get_connection()
    df   = pd.read_sql(query, conn)
    conn.close()
    return df


# ─── KPI Queries ───────────────────────────────────────────
QUERIES = {
    "Monthly Revenue": """
        SELECT
            DATE_FORMAT(t.transaction_date, '%Y-%m') AS Month,
            ROUND(SUM(CASE WHEN tc.category_type = 'REVENUE' AND t.transaction_type = 'CREDIT'
                           THEN t.amount ELSE 0 END), 2) AS Revenue,
            ROUND(SUM(CASE WHEN tc.category_type = 'COGS'    AND t.transaction_type = 'DEBIT'
                           THEN t.amount ELSE 0 END), 2) AS COGS,
            ROUND(SUM(CASE WHEN tc.category_type = 'EXPENSE' AND t.transaction_type = 'DEBIT'
                           THEN t.amount ELSE 0 END), 2) AS Expenses
        FROM transactions t
        JOIN accounts a             ON t.account_id  = a.account_id
        JOIN transaction_categories tc ON a.category_id = tc.category_id
        GROUP BY DATE_FORMAT(t.transaction_date, '%Y-%m')
        ORDER BY Month
    """,
    "Gross Margin": """
        WITH m AS (
            SELECT
                DATE_FORMAT(t.transaction_date, '%Y-%m') AS Month,
                SUM(CASE WHEN tc.category_type = 'REVENUE' AND t.transaction_type = 'CREDIT'
                         THEN t.amount ELSE 0 END) AS Revenue,
                SUM(CASE WHEN tc.category_type = 'COGS'    AND t.transaction_type = 'DEBIT'
                         THEN t.amount ELSE 0 END) AS COGS
            FROM transactions t
            JOIN accounts a             ON t.account_id  = a.account_id
            JOIN transaction_categories tc ON a.category_id = tc.category_id
            GROUP BY DATE_FORMAT(t.transaction_date, '%Y-%m')
        )
        SELECT Month, Revenue, COGS,
               ROUND(Revenue - COGS, 2) AS Gross_Profit,
               ROUND((Revenue - COGS) / NULLIF(Revenue, 0) * 100, 2) AS Gross_Margin_Pct
        FROM m ORDER BY Month
    """,
    "Budget vs Actual": """
        SELECT d.department_name AS Department,
               b.fiscal_month   AS Month,
               ROUND(COALESCE(SUM(t.amount), 0), 2) AS Actual,
               ROUND(SUM(b.budget_amount), 2)         AS Budget,
               ROUND(COALESCE(SUM(t.amount), 0) - SUM(b.budget_amount), 2) AS Variance
        FROM budgets b
        JOIN departments d ON b.department_id = d.department_id
        LEFT JOIN transactions t
            ON t.department_id = b.department_id
           AND MONTH(t.transaction_date) = b.fiscal_month
           AND YEAR(t.transaction_date)  = b.fiscal_year
        WHERE b.fiscal_year = 2024
        GROUP BY d.department_name, b.fiscal_month
        ORDER BY Department, Month
    """
}


def style_header_row(ws, row: int, num_cols: int, bg_color: str = NAVY) -> None:
    """Apply professional header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, col=col if col else 1)
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill      = PatternFill("solid", fgColor=bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws) -> None:
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def write_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    """Write a DataFrame to a styled Excel sheet."""
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    wb = writer.book
    ws = writer.sheets[sheet_name]

    # Title row
    ws.cell(row=1, column=1).value = sheet_name
    ws.cell(row=1, column=1).font  = Font(bold=True, size=14, color=NAVY, name="Calibri")

    # Header row (row 2)
    style_header_row(ws, row=2, num_cols=len(df.columns))

    # Zebra stripe data rows
    light_fill = PatternFill("solid", fgColor="F0F4FF")
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=0):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = light_fill
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(name="Calibri", size=10)

    auto_width(ws)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, REPORT_NAME)

    print("📊 Financial KPI Report Generator")
    print("=" * 40)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, query in QUERIES.items():
            print(f"  Fetching: {sheet_name}...")
            df = fetch_df(query)
            write_sheet(writer, sheet_name, df)
            print(f"  ✅ {len(df)} rows → '{sheet_name}' sheet")

    print(f"\n✅ Report saved: {output_path}")
    print("Open in Excel or Power BI for visualization.")


if __name__ == "__main__":
    main()
